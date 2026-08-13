from __future__ import annotations

import csv
import io
import json
import re
import uuid
from pathlib import Path
from typing import Any

from fastapi import APIRouter, File, Form, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
import pdfplumber
from pydantic import BaseModel

from app.repositories.yp_data_repository import YPDataRepository


router = APIRouter(prefix="/yp", tags=["carrier-data"])
repository = YPDataRepository()
STAGE_DIR = Path(__file__).resolve().parents[2] / "YP_upload_stages"

FIELDS = ["mode", "origin_name", "destination_name", "typical_base_rate", "typical_transit_hours", "capacity_value", "capacity_unit", "on_time_rate"]
REQUIRED = {"mode", "origin_name", "destination_name"}
ALIASES = {
    "transport mode": "mode", "운송수단": "mode", "수단": "mode",
    "origin": "origin_name", "origin name": "origin_name", "출발지": "origin_name", "출발지명": "origin_name",
    "destination": "destination_name", "destination name": "destination_name", "도착지": "destination_name", "도착지명": "destination_name",
    "all in rate": "typical_base_rate", "rate": "typical_base_rate", "운임": "typical_base_rate", "기본 운임": "typical_base_rate",
    "transit hours": "typical_transit_hours", "소요시간": "typical_transit_hours",
    "capacity": "capacity_value", "수송량": "capacity_value", "capacity unit": "capacity_unit", "수송단위": "capacity_unit",
    "on time rate": "on_time_rate", "정시율": "on_time_rate",
}


class RemapRequest(BaseModel):
    mappings: dict[str, str | None]


def stage_path(upload_id: str) -> Path:
    return STAGE_DIR / f"YP_{upload_id}.json"


def save_stage(stage: dict[str, Any]) -> None:
    STAGE_DIR.mkdir(exist_ok=True)
    stage_path(stage["upload_id"]).write_text(json.dumps(stage, ensure_ascii=False), encoding="utf-8")


def load_stage(upload_id: str) -> dict[str, Any]:
    path = stage_path(upload_id)
    if not path.exists():
        raise HTTPException(status_code=404, detail="업로드 분석 결과를 찾을 수 없습니다.")
    return json.loads(path.read_text(encoding="utf-8"))


@router.get("/uploads/template")
def template() -> StreamingResponse:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "carrier_capabilities"
    sheet.append(FIELDS)
    sheet.append(["sea", "Busan", "Hamburg", 1200, 360, 100, "vehicle", 0.92])
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": 'attachment; filename="glovis_carrier_capability_template.xlsx"'})


def read_rows(filename: str, content: bytes) -> tuple[list[str], list[dict[str, Any]]]:
    suffix = filename.lower().rsplit(".", 1)[-1]
    if suffix == "csv":
        text = content.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text))
        return list(reader.fieldnames or []), list(reader)
    if suffix in {"xlsx", "xlsm"}:
        sheet = load_workbook(io.BytesIO(content), read_only=True, data_only=True).active
        values = list(sheet.values)
        headers = [str(value or "").strip() for value in values[0]] if values else []
        return headers, [dict(zip(headers, row)) for row in values[1:] if any(value is not None for value in row)]
    if suffix == "pdf":
        tables: list[list[list[Any]]] = []
        with pdfplumber.open(io.BytesIO(content)) as pdf:
            for page in pdf.pages:
                tables.extend(table for table in (page.extract_tables() or []) if table and len(table) > 1)
        if not tables:
            raise HTTPException(status_code=422, detail="PDF에서 표를 찾지 못했습니다.")
        matrix = max(tables, key=lambda table: len(table) * len(table[0]))
        headers = [str(value or "").strip() for value in matrix[0]]
        return headers, [dict(zip(headers, row)) for row in matrix[1:] if any(value not in {None, ""} for value in row)]
    raise HTTPException(status_code=400, detail="CSV, XLSX, PDF 파일만 분석할 수 있습니다.")


def header_key(value: str) -> str:
    normalized = re.sub(r"[_\-]+", " ", value.strip().lower())
    return ALIASES.get(normalized, normalized.replace(" ", "_") if normalized.replace(" ", "_") in FIELDS else "")


def auto_mappings(headers: list[str]) -> dict[str, str | None]:
    mappings = {field: None for field in FIELDS}
    for source in headers:
        target = header_key(source)
        if target and mappings[target] is None:
            mappings[target] = source
    return mappings


def number(value: Any, field: str, errors: list[str]) -> float | None:
    if value in {None, ""}:
        return None
    try:
        return float(str(value).replace(",", "").replace("%", "").strip())
    except ValueError:
        errors.append(f"{field}는 숫자여야 합니다.")
        return None


def validate_rows(raw_rows: list[dict[str, Any]], mappings: dict[str, str | None]) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    valid: list[dict[str, Any]] = []
    issues: list[dict[str, Any]] = []
    for index, source in enumerate(raw_rows, start=2):
        row = {field: source.get(column) if column else None for field, column in mappings.items()}
        errors = [f"{field} 필수값 누락" for field in REQUIRED if not row.get(field)]
        mode = str(row.get("mode") or "").strip().lower()
        if mode == "truck": mode = "road"
        if mode and mode not in {"sea", "air", "rail", "road"}: errors.append("mode는 sea, air, rail, road 중 하나여야 합니다.")
        row["mode"] = mode
        for field in ("typical_base_rate", "typical_transit_hours", "capacity_value", "on_time_rate"):
            row[field] = number(row.get(field), field, errors)
        if row.get("on_time_rate") is not None and row["on_time_rate"] > 1:
            row["on_time_rate"] = row["on_time_rate"] / 100
        if row.get("on_time_rate") is not None and not 0 <= row["on_time_rate"] <= 1:
            errors.append("on_time_rate는 0~1 범위여야 합니다.")
        if errors:
            issues.append({"row": index, "errors": errors, "data_errors": errors, "warnings": [], "data_warnings": []})
        else:
            valid.append(row)
    return valid, issues


def analysis_payload(stage: dict[str, Any]) -> dict[str, Any]:
    return {
        "upload_id": stage["upload_id"], "file_name": stage["file_name"], "carrier_name": stage["carrier_name"],
        "total_rows": len(stage["raw_rows"]), "valid_rows": len(stage["rows"]), "invalid_rows": len(stage["raw_rows"]) - len(stage["rows"]),
        "warning_count": 0, "headers": stage["headers"],
        "mapping": [{"system_field": field, "source_column": stage["mappings"].get(field), "required": field in REQUIRED, "status": "mapped" if stage["mappings"].get(field) else "excluded"} for field in FIELDS],
        "issues": stage["issues"][:100], "preview": stage["rows"][:10],
    }


@router.post("/uploads/analyze")
async def analyze(file: UploadFile = File(...), carrier_name: str = Form(...)) -> dict[str, Any]:
    if not carrier_name.strip():
        raise HTTPException(status_code=400, detail="선사 이름을 입력해 주세요.")
    content = await file.read()
    headers, rows = read_rows(file.filename or "", content)
    mappings = auto_mappings(headers)
    valid_rows, issues = validate_rows(rows, mappings)
    upload_id = uuid.uuid4().hex
    stage = {"upload_id": upload_id, "carrier_name": carrier_name.strip(), "rows": valid_rows, "raw_rows": rows, "headers": headers, "mappings": mappings, "issues": issues, "file_name": file.filename or "upload"}
    save_stage(stage)
    return analysis_payload(stage)


@router.post("/uploads/{upload_id}/remap")
def remap(upload_id: str, payload: RemapRequest) -> dict[str, Any]:
    stage = load_stage(upload_id)
    stage["mappings"] = {field: payload.mappings.get(field) for field in FIELDS}
    stage["rows"], stage["issues"] = validate_rows(stage["raw_rows"], stage["mappings"])
    save_stage(stage)
    return analysis_payload(stage)


@router.post("/uploads/{upload_id}/commit")
def commit(upload_id: str) -> dict[str, int]:
    upload = load_stage(upload_id)
    if not upload["rows"]:
        raise HTTPException(status_code=409, detail="DB에 반영할 유효 행이 없습니다.")
    inserted = repository.insert_upload(upload["carrier_name"], upload["rows"], upload["file_name"])
    stage_path(upload_id).unlink(missing_ok=True)
    return {"inserted_or_updated": inserted}
